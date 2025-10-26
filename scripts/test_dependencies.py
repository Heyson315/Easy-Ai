"""
Test script to verify all dependencies are working correctly
"""

def test_imports():
    """Test that all major packages can be imported"""
    print("🧪 Testing Rahman Analytics Hub Dependencies...")
    
    try:
        # Core Excel and Data Processing
        import openpyxl
        import pandas as pd
        import numpy as np
        print("✅ Core data processing packages: OK")
        
        # Data Science and Analytics
        import matplotlib.pyplot as plt
        import seaborn as sns
        import scipy
        from sklearn.model_selection import train_test_split
        print("✅ Data science packages: OK")
        
        # Visualization
        import plotly.graph_objects as go
        print("✅ Visualization packages: OK")
        
        # Development tools
        import pytest
        import jupyter
        print("✅ Development tools: OK")
        
        # Data validation
        import pydantic
        import jsonschema
        print("✅ Data validation packages: OK")
        
        print("\n🎉 All dependencies successfully imported!")
        return True
        
    except ImportError as e:
        print(f"❌ Import error: {e}")
        return False

def test_basic_functionality():
    """Test basic functionality with sample data"""
    print("\n🔬 Testing Basic Functionality...")
    
    try:
        import pandas as pd
        import numpy as np
        
        # Create sample data
        data = {
            'Date': pd.date_range('2025-01-01', periods=10),
            'Transaction': ['Initial Budget', 'Office Supplies', 'Marketing', 'Development'] * 2 + ['Training', 'Equipment'],
            'Amount': np.random.randint(100, 5000, 10)
        }
        df = pd.DataFrame(data)
        
        # Basic operations
        total = df['Amount'].sum()
        mean_amount = df['Amount'].mean()
        
        print(f"✅ Created sample dataset with {len(df)} rows")
        print(f"✅ Total amount: ${total:,}")
        print(f"✅ Average amount: ${mean_amount:.2f}")
        
        return True
        
    except Exception as e:
        print(f"❌ Functionality error: {e}")
        return False

def test_excel_generation():
    """Test Excel generation capabilities"""
    print("\n📊 Testing Excel Generation...")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # Create a test workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dependency Test"
        
        # Add headers
        headers = ["Package", "Version", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        
        # Add sample data
        packages = [
            ["openpyxl", "3.1.5", "✅ Working"],
            ["pandas", "2.3.3", "✅ Working"],
            ["numpy", "2.3.4", "✅ Working"]
        ]
        
        for row, package_info in enumerate(packages, 2):
            for col, value in enumerate(package_info, 1):
                ws.cell(row=row, column=col).value = value
        
        # Save test file
        test_file = "output/reports/financial/dependency_test.xlsx"
        wb.save(test_file)
        
        print(f"✅ Excel test file created: {test_file}")
        return True
        
    except Exception as e:
        print(f"❌ Excel generation error: {e}")
        return False

if __name__ == "__main__":
    print("🚀 Rahman Analytics Hub - Dependency Verification")
    print("=" * 50)
    
    # Run all tests
    import_success = test_imports()
    functionality_success = test_basic_functionality()
    excel_success = test_excel_generation()
    
    print("\n" + "=" * 50)
    if import_success and functionality_success and excel_success:
        print("🎉 ALL TESTS PASSED! Your environment is ready for development!")
        print("\n🚀 Next steps:")
        print("1. ✅ Dependencies installed and working")
        print("2. ✅ Excel generation tested")
        print("3. ✅ Data processing capabilities verified")
        print("4. 📈 Ready to build your first analytics pipeline!")
    else:
        print("⚠️  Some tests failed. Please check the error messages above.")