"""
SharePoint permissions data processor
Analyzes the CSV data from SharePoint for compliance and security insights
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import json

def load_sharepoint_data():
    """Load SharePoint permissions data from CSV"""
    data_path = "data/raw/sharepoint/Hassan Rahman_2025-8-16-20-24-4_1.csv"
    
    try:
        df = pd.read_csv(data_path)
        print(f"✅ Loaded {len(df)} SharePoint permission records")
        return df
    except Exception as e:
        print(f"❌ Error loading data: {e}")
        return None

def analyze_permissions(df):
    """Analyze SharePoint permissions for insights"""
    
    # Basic statistics
    analysis = {
        'total_permissions': len(df),
        'unique_users': df['User Name'].nunique(),
        'unique_resources': df['Resource Path'].nunique(),
        'permission_types': df['Permission'].value_counts().to_dict(),
        'user_or_group_types': df['User Or Group Type'].value_counts().to_dict()
    }
    
    return analysis

def create_permissions_report(df, analysis):
    """Create Excel report of SharePoint permissions analysis"""
    
    wb = openpyxl.Workbook()
    
    # Summary sheet
    summary_sheet = wb.active
    summary_sheet.title = "Permissions Summary"
    
    # Headers
    headers = ["Metric", "Value", "Details"]
    for col, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    
    # Summary data
    summary_data = [
        ["Total Permissions", analysis['total_permissions'], "Total permission entries"],
        ["Unique Users", analysis['unique_users'], "Number of distinct users"],
        ["Unique Resources", analysis['unique_resources'], "Number of distinct resources"],
        ["Most Common Permission", max(analysis['permission_types'], key=analysis['permission_types'].get), "Most frequently granted permission"]
    ]
    
    for row, data in enumerate(summary_data, 2):
        for col, value in enumerate(data, 1):
            summary_sheet.cell(row=row, column=col).value = value
    
    # Raw data sheet
    raw_sheet = wb.create_sheet(title="Raw Permissions Data")
    
    # Add raw data to sheet
    raw_headers = list(df.columns)
    for col, header in enumerate(raw_headers, 1):
        cell = raw_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    
    # Add first 100 rows of data
    for row_idx, row_data in enumerate(df.head(100).itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            raw_sheet.cell(row=row_idx, column=col_idx).value = str(value)
    
    # Save report
    output_path = f"output/reports/business/SharePoint_Permissions_Analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(output_path)
    print(f"📊 Report saved: {output_path}")
    
    return output_path

def main():
    """Main analysis workflow"""
    print("🔐 SharePoint Permissions Analysis")
    print("=" * 40)
    
    # Load data
    df = load_sharepoint_data()
    if df is None:
        return
    
    # Analyze
    analysis = analyze_permissions(df)
    
    # Display key insights
    print(f"\n📊 Key Insights:")
    print(f"• Total permission entries: {analysis['total_permissions']:,}")
    print(f"• Unique users: {analysis['unique_users']:,}")
    print(f"• Unique resources: {analysis['unique_resources']:,}")
    print(f"• Most common permission: {max(analysis['permission_types'], key=analysis['permission_types'].get)}")
    
    # Create report
    report_path = create_permissions_report(df, analysis)
    
    print(f"\n✅ Analysis complete! Check: {report_path}")

if __name__ == "__main__":
    main()